"""
Generate consolidated_metadata_with_descriptions.xlsx for the chatbot's
system context.

Produces three sheets matching the reference format:
  - Column Descriptions   (COLUMN_NAME, COLUMN_DESCRIPTION, COLUMN_DATA_TYPE,
                           SAMPLE_VALUES, TABLE_NAME)
  - Table Descriptions    (TABLE_NAME, TABLE_DESCRIPTION)
  - Table Relationships   (TABLE_NAME, COLUMN_NAME, RELATED_TABLE_NAME,
                           RELATED_COLUMN_NAME)

Descriptions are curated (deterministic, no LLM round-trips needed) since the
schema is small and stable. Sample values come from the latest parquet file
per table.

Run:
    python -m metadata          # writes to ../consolidated_metadata_with_descriptions.xlsx
    python -m metadata <path>   # writes to <path>
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import storage

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Curated table & column descriptions
# ---------------------------------------------------------------------------
# Hand-written so the LLM always gets stable, accurate, business-meaningful
# context. Update whenever models.py changes.

TABLE_DESCRIPTIONS: dict[str, str] = {
    "billing_usage": (
        "Per-record Databricks billing events from system.billing.usage. One row per "
        "compute usage interval (cluster run, warehouse query batch, job execution). "
        "Joined to list_prices on (sku_name, cloud, usage_unit, time range) to compute cost. "
        "usage_usd is pre-calculated when available; otherwise multiply usage_quantity x "
        "list_prices.effective_list_price. Filter by usage_date for time ranges."
    ),
    "list_prices": (
        "Effective DBU prices per SKU/cloud/usage_unit from system.billing.list_prices. "
        "Each row covers a time range [price_start_time, price_end_time); price_end_time NULL "
        "means currently active. Use effective_list_price (negotiated) over default_price."
    ),
    "clusters": (
        "Cluster configuration history from system.compute.clusters. One row per "
        "configuration change event for each cluster_id, so the same cluster_id can appear "
        "multiple times. To get the latest config use DISTINCT ON (cluster_id) ORDER BY "
        "change_time DESC. Join to billing_usage on cluster_id."
    ),
    "warehouses": (
        "SQL warehouse configuration history from system.compute.warehouses. One row per "
        "configuration change event per warehouse_id. created_by is the user who made the "
        "change. Join to billing_usage on warehouse_id. Note: billing_usage.run_as is "
        "rarely populated for warehouse rows in the source system table."
    ),
    "jobs": (
        "Job definitions from system.lakeflow.jobs. One row per job (or per change event). "
        "creator_id and run_as identify ownership/execution identity. Join to billing_usage "
        "on job_id."
    ),
}

COLUMN_DESCRIPTIONS: dict[tuple[str, str], str] = {
    # billing_usage
    ("billing_usage", "id"): "Synthetic primary key (autoincrement).",
    ("billing_usage", "account_id"): "Databricks account identifier.",
    ("billing_usage", "workspace_id"): "Workspace where the usage occurred. Numeric string.",
    ("billing_usage", "record_id"): "Unique identifier for this usage record (UUID).",
    ("billing_usage", "sku_name"): (
        "SKU name identifying the compute product / pricing tier "
        "(e.g. PREMIUM_JOBS_COMPUTE, SERVERLESS_SQL_COMPUTE)."
    ),
    ("billing_usage", "cloud"): "Cloud provider: AZURE, AWS, or GCP.",
    ("billing_usage", "usage_start_time"): "Timestamp when the usage interval started.",
    ("billing_usage", "usage_end_time"): "Timestamp when the usage interval ended.",
    ("billing_usage", "usage_date"): "Calendar date of the usage (use this for date-range filters).",
    ("billing_usage", "usage_unit"): "Unit of measurement (typically 'DBU').",
    ("billing_usage", "usage_quantity"): "Quantity consumed in usage_unit (i.e. DBUs consumed).",
    ("billing_usage", "billing_origin_product"): (
        "Product that originated the usage: JOBS, SQL, ALL_PURPOSE, DLT, MODEL_SERVING, SERVERLESS."
    ),
    ("billing_usage", "usage_type"): (
        "How compute was consumed: COMPUTE_TIME, STORAGE_SPACE, NETWORK_BYTES, TOKEN, GPU_TIME."
    ),
    ("billing_usage", "record_type"): "ORIGINAL or CORRECTION.",
    ("billing_usage", "ingestion_date"): "Date this row was ingested into the source table.",
    ("billing_usage", "cluster_id"): "Cluster the usage came from. NULL for non-cluster usage. Joins to clusters.cluster_id.",
    ("billing_usage", "warehouse_id"): "SQL warehouse the usage came from. NULL for non-warehouse usage. Joins to warehouses.warehouse_id.",
    ("billing_usage", "node_type"): "Cloud VM type for the underlying compute (e.g. Standard_DS3_v2).",
    ("billing_usage", "job_id"): "Job that drove the usage. NULL if not a job. Joins to jobs.job_id.",
    ("billing_usage", "run_name"): "Human-readable name of the job run.",
    ("billing_usage", "run_as"): (
        "Identity that executed the workload. Populated for cluster/job rows; rarely populated "
        "for warehouse rows (Databricks data quirk)."
    ),
    ("billing_usage", "jobs_tier"): "STANDARD/PREMIUM/ENTERPRISE for JOBS-origin rows.",
    ("billing_usage", "sql_tier"): "STANDARD/PREMIUM/ENTERPRISE for SQL/SERVERLESS-origin rows.",
    ("billing_usage", "dlt_tier"): "STANDARD/PREMIUM/ENTERPRISE for DLT-origin rows.",
    ("billing_usage", "is_serverless"): "True if the usage ran on serverless compute.",
    ("billing_usage", "is_photon"): "True if the Photon engine was used.",
    ("billing_usage", "serving_type"): "Model-serving type (e.g. MODEL_SERVING) when applicable.",
    ("billing_usage", "instance_pool_id"): "Instance pool ID if the cluster was launched from a pool.",
    ("billing_usage", "usage_usd"): (
        "Pre-calculated cost in USD (usage_quantity x effective_list_price). Use this when "
        "available; otherwise join to list_prices."
    ),

    # list_prices
    ("list_prices", "id"): "Synthetic primary key.",
    ("list_prices", "account_id"): "Databricks account identifier.",
    ("list_prices", "sku_name"): "SKU name; matches billing_usage.sku_name.",
    ("list_prices", "cloud"): "Cloud provider; matches billing_usage.cloud.",
    ("list_prices", "currency_code"): "ISO currency code (typically USD).",
    ("list_prices", "usage_unit"): "Unit of pricing (e.g. DBU); matches billing_usage.usage_unit.",
    ("list_prices", "price_start_time"): "Inclusive start of the price's effective window.",
    ("list_prices", "price_end_time"): "Exclusive end of the window. NULL = currently active.",
    ("list_prices", "default_price"): "List/rack price per usage_unit.",
    ("list_prices", "effective_list_price"): "Price per usage_unit after negotiated discount. Use this for cost calc.",

    # clusters
    ("clusters", "id"): "Synthetic primary key.",
    ("clusters", "account_id"): "Databricks account identifier.",
    ("clusters", "workspace_id"): "Workspace the cluster belongs to.",
    ("clusters", "cluster_id"): "Cluster identifier; joins to billing_usage.cluster_id.",
    ("clusters", "cluster_name"): "Human-readable cluster name.",
    ("clusters", "owned_by"): "Email of the cluster owner.",
    ("clusters", "driver_node_type"): "VM type for the driver node.",
    ("clusters", "worker_node_type"): "VM type for worker nodes.",
    ("clusters", "worker_count"): "Fixed worker count for non-autoscaling clusters.",
    ("clusters", "min_autoscale_workers"): "Lower bound for autoscaling.",
    ("clusters", "max_autoscale_workers"): "Upper bound for autoscaling.",
    ("clusters", "dbr_version"): "Databricks Runtime version (e.g. 14.3.x-scala2.12).",
    ("clusters", "cluster_source"): "How it was created: JOB, UI, PIPELINE, PIPELINE_MAINTENANCE.",
    ("clusters", "data_security_mode"): "Access control mode (SINGLE_USER, USER_ISOLATION, NONE).",
    ("clusters", "create_time"): "When the cluster was created.",
    ("clusters", "delete_time"): "When the cluster was deleted (NULL if still active).",
    ("clusters", "change_time"): "When this configuration row was written. Latest row = latest config.",

    # warehouses
    ("warehouses", "id"): "Synthetic primary key.",
    ("warehouses", "account_id"): "Databricks account identifier.",
    ("warehouses", "workspace_id"): "Workspace the warehouse belongs to.",
    ("warehouses", "warehouse_id"): "Warehouse identifier; joins to billing_usage.warehouse_id.",
    ("warehouses", "warehouse_name"): "Human-readable warehouse name.",
    ("warehouses", "warehouse_type"): "CLASSIC, PRO, or SERVERLESS.",
    ("warehouses", "warehouse_size"): "T-shirt size: 2X_SMALL, X_SMALL, SMALL, MEDIUM, LARGE, X_LARGE, 2X_LARGE, 3X_LARGE, 4X_LARGE.",
    ("warehouses", "min_clusters"): "Lower bound on concurrent cluster scaling.",
    ("warehouses", "max_clusters"): "Upper bound on concurrent cluster scaling.",
    ("warehouses", "auto_stop_minutes"): "Idle minutes before auto-stop.",
    ("warehouses", "created_by"): (
        "User who created the warehouse. Use this (not billing_usage.run_as) for per-user "
        "warehouse cost attribution."
    ),
    ("warehouses", "change_time"): "When this configuration row was written.",
    ("warehouses", "delete_time"): "When the warehouse was deleted (NULL if still active).",

    # jobs
    ("jobs", "id"): "Synthetic primary key.",
    ("jobs", "account_id"): "Databricks account identifier.",
    ("jobs", "workspace_id"): "Workspace the job belongs to.",
    ("jobs", "job_id"): "Job identifier; joins to billing_usage.job_id.",
    ("jobs", "name"): "Human-readable job name.",
    ("jobs", "creator_id"): "User who created the job.",
    ("jobs", "run_as"): "Identity the job runs as.",
    ("jobs", "change_time"): "When this configuration row was written.",
    ("jobs", "delete_time"): "When the job was deleted (NULL if still active).",
}

# Foreign-key style relationships across tables
RELATIONSHIPS: list[tuple[str, str, str, str]] = [
    ("billing_usage", "sku_name",     "list_prices", "sku_name"),
    ("billing_usage", "cloud",        "list_prices", "cloud"),
    ("billing_usage", "usage_unit",   "list_prices", "usage_unit"),
    ("billing_usage", "cluster_id",   "clusters",    "cluster_id"),
    ("billing_usage", "warehouse_id", "warehouses",  "warehouse_id"),
    ("billing_usage", "job_id",       "jobs",        "job_id"),
    ("billing_usage", "workspace_id", "clusters",    "workspace_id"),
    ("billing_usage", "workspace_id", "warehouses",  "workspace_id"),
    ("billing_usage", "workspace_id", "jobs",        "workspace_id"),
    ("billing_usage", "account_id",   "list_prices", "account_id"),
    ("billing_usage", "account_id",   "clusters",    "account_id"),
    ("billing_usage", "account_id",   "warehouses",  "account_id"),
    ("billing_usage", "account_id",   "jobs",        "account_id"),
]


def _latest_parquet(data_dir: Path, table: str) -> Path | str | None:
    """Find the most recent parquet for ``table``.

    For backward compatibility this still accepts a ``data_dir`` argument
    (used when called as a script with an explicit dir). When the configured
    DATA_STORE is *not* local, ``data_dir`` is ignored and the storage
    layer's URI is returned (s3:// / az:// / gs://).
    """
    if storage.backing_store() != "local":
        return storage.latest_parquet(table)
    matches = sorted(data_dir.glob(f"{table}_*.parquet"))
    return matches[-1] if matches else None


def _sample_values(df: pd.DataFrame, col: str, n: int = 5) -> list[Any]:
    """Up to N distinct non-null sample values from a column, JSON-safe."""
    series = df[col].dropna() if col in df.columns else pd.Series(dtype=object)
    if series.empty:
        return []
    uniq = series.drop_duplicates().head(n).tolist()
    out: list[Any] = []
    for v in uniq:
        # Make timestamps/dates string-friendly
        if hasattr(v, "isoformat"):
            out.append(v.isoformat())
        else:
            out.append(v)
    return out


def _spark_like_dtype(dtype: Any) -> str:
    """Render a pandas/numpy dtype in a readable form."""
    s = str(dtype)
    mapping = {
        "object": "StringType()",
        "int64": "LongType()",
        "int32": "IntegerType()",
        "float64": "DoubleType()",
        "float32": "FloatType()",
        "bool": "BooleanType()",
        "datetime64[ns]": "TimestampType()",
        "datetime64[ns, UTC]": "TimestampType()",
    }
    return mapping.get(s, s)


def build_metadata(data_dir: Path, output_path: Path) -> Path:
    """Build the consolidated metadata workbook from parquet files + curated text."""
    table_order = ["billing_usage", "list_prices", "clusters", "warehouses", "jobs"]

    column_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []

    for table in table_order:
        pq = _latest_parquet(data_dir, table)
        if pq is None:
            logger.warning("No parquet file found for %s in %s; skipping samples", table, data_dir)
            df = pd.DataFrame()
        else:
            df = storage.read_parquet(str(pq))
            display_name = pq.name if isinstance(pq, Path) else str(pq).rsplit("/", 1)[-1]
            logger.info("Loaded %s rows from %s", len(df), display_name)

        # Table description row
        table_rows.append({
            "TABLE_NAME": table,
            "TABLE_DESCRIPTION": TABLE_DESCRIPTIONS.get(table, ""),
        })

        # Column rows (use parquet schema as source of truth; fall back to curated dict)
        col_names = list(df.columns) if not df.empty else [
            c for (t, c) in COLUMN_DESCRIPTIONS.keys() if t == table
        ]
        for col in col_names:
            dtype = df[col].dtype if not df.empty and col in df.columns else "object"
            column_rows.append({
                "COLUMN_NAME": col,
                "COLUMN_DESCRIPTION": COLUMN_DESCRIPTIONS.get((table, col), ""),
                "COLUMN_DATA_TYPE": _spark_like_dtype(dtype),
                "SAMPLE_VALUES": str(_sample_values(df, col)) if not df.empty else "[]",
                "TABLE_NAME": table,
            })

    # Relationships
    rel_rows = [
        {
            "TABLE_NAME": t,
            "COLUMN_NAME": c,
            "RELATED_TABLE_NAME": rt,
            "RELATED_COLUMN_NAME": rc,
        }
        for (t, c, rt, rc) in RELATIONSHIPS
    ]

    # Write workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # The default sheet
    default = wb.active
    default.title = "Column Descriptions"
    for r in dataframe_to_rows(pd.DataFrame(column_rows), index=False, header=True):
        default.append(r)

    table_ws = wb.create_sheet("Table Descriptions")
    for r in dataframe_to_rows(pd.DataFrame(table_rows), index=False, header=True):
        table_ws.append(r)

    rel_ws = wb.create_sheet("Table Relationships")
    for r in dataframe_to_rows(pd.DataFrame(rel_rows), index=False, header=True):
        rel_ws.append(r)

    wb.save(output_path)
    logger.info("Wrote metadata workbook to %s", output_path)
    return output_path


def main(argv: list[str]) -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    backend_dir = Path(__file__).resolve().parent
    data_dir = backend_dir.parent / "data"
    if argv:
        out = Path(argv[0]).resolve()
    else:
        out = backend_dir.parent / "consolidated_metadata_with_descriptions.xlsx"
    build_metadata(data_dir, out)
    print(f"OK -> {out}")


if __name__ == "__main__":
    main(sys.argv[1:])
